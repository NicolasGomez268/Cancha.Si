from django.template.loader import render_to_string
from django.utils import timezone
from datetime import datetime, timedelta
import os

try:
    from weasyprint import HTML
    WEASYPRINT_ENABLED = True
except OSError:
    WEASYPRINT_ENABLED = False
    print("WeasyPrint no está disponible. Los reportes PDF no funcionarán.")

from openpyxl import Workbook
from .models import Complejo, Cancha, Reserva

class ReporteService:
    @staticmethod
    def generar_reporte_complejo(complejo, fecha_inicio=None, fecha_fin=None):
        """
        Genera un reporte PDF para un complejo deportivo en un rango de fechas.
        Si no se especifican fechas, genera el reporte del último mes.
        """
        if not WEASYPRINT_ENABLED:
            return None

        # Si no se especifican fechas, usar el último mes
        if not fecha_fin:
            fecha_fin = timezone.now().date()
        if not fecha_inicio:
            fecha_inicio = fecha_fin - timedelta(days=30)

        # Obtener datos para el reporte
        reservas = complejo.reservas.filter(
            fecha_hora__date__range=[fecha_inicio, fecha_fin],
            cancelada=False
        ).select_related('cancha', 'jugador')

        # Calcular estadísticas
        total_reservas = reservas.count()
        ingresos_totales = sum(r.monto_total for r in reservas if r.pagado)
        canchas_mas_reservadas = complejo.canchas.all().order_by('-reservas__count')[:5]
        
        # Preparar datos por día
        datos_diarios = {}
        for reserva in reservas:
            fecha = reserva.fecha_hora.date()
            if fecha not in datos_diarios:
                datos_diarios[fecha] = {
                    'total_reservas': 0,
                    'ingresos': 0,
                }
            datos_diarios[fecha]['total_reservas'] += 1
            if reserva.pagado:
                datos_diarios[fecha]['ingresos'] += reserva.monto_total

        # Preparar contexto para el template
        context = {
            'complejo': complejo,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'fecha_generacion': timezone.now(),
            'total_reservas': total_reservas,
            'ingresos_totales': ingresos_totales,
            'promedio_diario': ingresos_totales / (fecha_fin - fecha_inicio).days if total_reservas > 0 else 0,
            'canchas_mas_reservadas': canchas_mas_reservadas,
            'datos_diarios': sorted(datos_diarios.items()),
            'reservas': reservas,
        }

        # Renderizar el HTML
        html_string = render_to_string('canchas/reportes/reporte_complejo.html', context)

        # Crear el PDF
        html = HTML(string=html_string)
        
        # Crear directorio para reportes si no existe
        reportes_dir = os.path.join('media', 'reportes')
        os.makedirs(reportes_dir, exist_ok=True)
        
        # Generar nombre de archivo
        nombre_archivo = f'reporte_{complejo.id}_{fecha_inicio.strftime("%Y%m%d")}_{fecha_fin.strftime("%Y%m%d")}.pdf'
        ruta_archivo = os.path.join(reportes_dir, nombre_archivo)
        
        # Guardar PDF
        html.write_pdf(ruta_archivo)
        
        return ruta_archivo

    @staticmethod
    def generar_reporte_reserva(reserva):
        """
        Genera un PDF con los detalles de una reserva específica
        """
        if not WEASYPRINT_ENABLED:
            return None

        context = {
            'reserva': reserva,
            'fecha_generacion': timezone.now(),
        }

        html_string = render_to_string('canchas/reportes/comprobante_reserva.html', context)
        html = HTML(string=html_string)
        
        reportes_dir = os.path.join('media', 'reportes', 'comprobantes')
        os.makedirs(reportes_dir, exist_ok=True)
        
        nombre_archivo = f'comprobante_reserva_{reserva.id}.pdf'
        ruta_archivo = os.path.join(reportes_dir, nombre_archivo)
        
        html.write_pdf(ruta_archivo)
        return ruta_archivo

class EstadisticasService:
    @staticmethod
    def obtener_estadisticas_complejo(complejo, fecha_inicio=None, fecha_fin=None):
        """
        Obtiene estadísticas detalladas de un complejo para un rango de fechas
        """
        if not fecha_fin:
            fecha_fin = timezone.now().date()
        if not fecha_inicio:
            fecha_inicio = fecha_fin - timedelta(days=30)

        reservas = complejo.reservas.filter(
            fecha_hora__date__range=[fecha_inicio, fecha_fin]
        )

        return {
            'total_reservas': reservas.count(),
            'ingresos_totales': sum(r.monto_total for r in reservas if r.pagado),
            'reservas_canceladas': reservas.filter(cancelada=True).count(),
            'tasa_ocupacion': complejo.calcular_tasa_ocupacion(fecha_inicio, fecha_fin),
            'horarios_populares': complejo.obtener_horarios_populares(),
            'canchas_mas_reservadas': complejo.obtener_canchas_mas_reservadas(),
        }

def exportar_estadisticas_complejo_excel(complejo_id):
    """
    Exporta las estadísticas de un complejo a Excel
    """
    complejo = Complejo.objects.get(id=complejo_id)
    wb = Workbook()
    
    # Hoja de Resumen
    ws = wb.active
    ws.title = "Resumen"
    
    # Información del Complejo
    ws['A1'] = 'Estadísticas del Complejo'
    ws['A2'] = 'Nombre:'
    ws['B2'] = complejo.nombre
    ws['A3'] = 'Dueño:'
    ws['B3'] = complejo.dueno.usuario.get_full_name()
    ws['A4'] = 'Teléfono:'
    ws['B4'] = complejo.telefono
    ws['A5'] = 'Stock Cantina:'
    ws['B5'] = complejo.stock_cantina
    ws['A6'] = 'Ingresos Mensuales:'
    ws['B6'] = f'${complejo.ingresos_mensuales}'
    
    # Estadísticas Generales
    ws['A8'] = 'Estadísticas Generales'
    ws['A9'] = 'Total Canchas:'
    ws['B9'] = complejo.canchas.count()
    ws['A10'] = 'Total Reservas:'
    total_reservas = Reserva.objects.filter(cancha__complejo=complejo).count()
    ws['B10'] = total_reservas
    ws['A11'] = 'Reservas Activas:'
    reservas_activas = Reserva.objects.filter(cancha__complejo=complejo, cancelada=False).count()
    ws['B11'] = reservas_activas
    ws['A12'] = 'Reservas Canceladas:'
    ws['B12'] = total_reservas - reservas_activas
    
    # Detalles de Canchas
    ws_canchas = wb.create_sheet("Canchas")
    ws_canchas.append(['Nombre', 'Precio por Hora', 'Servicios', 'Total Reservas', 'Ingresos Generados'])
    
    for cancha in complejo.canchas.all():
        reservas_cancha = Reserva.objects.filter(cancha=cancha, cancelada=False)
        total_reservas = reservas_cancha.count()
        ingresos_cancha = sum(r.precio_total for r in reservas_cancha)
        ws_canchas.append([
            cancha.nombre,
            cancha.precio_hora,
            cancha.servicios,
            total_reservas,
            f'${ingresos_cancha}'
        ])
    
    # Detalles de Reservas
    ws_reservas = wb.create_sheet("Reservas")
    ws_reservas.append(['Fecha', 'Cancha', 'Jugador', 'Precio Total', 'Estado', 'Seña Pagada'])
    
    reservas = Reserva.objects.filter(cancha__complejo=complejo)
    for reserva in reservas:
        estado = 'Cancelada' if reserva.cancelada else 'Activa'
        ws_reservas.append([
            reserva.fecha_hora.strftime('%d/%m/%Y %H:%M'),
            reserva.cancha.nombre,
            reserva.jugador.get_full_name(),
            f'${reserva.precio_total}',
            estado,
            'Sí' if reserva.sena_pagada else 'No'
        ])
    
    # Estadísticas por Mes
    ws_mensual = wb.create_sheet("Estadísticas Mensuales")
    ws_mensual.append(['Mes', 'Total Reservas', 'Ingresos', 'Cancelaciones'])
    
    from django.db.models import Count
    from django.db.models.functions import TruncMonth
    
    stats_mensuales = (
        Reserva.objects
        .filter(cancha__complejo=complejo)
        .annotate(mes=TruncMonth('fecha_hora'))
        .values('mes')
        .annotate(
            total=Count('id'),
            canceladas=Count('id', filter=models.Q(cancelada=True)),
            ingresos=models.Sum('precio_total', filter=models.Q(cancelada=False))
        )
        .order_by('mes')
    )
    
    for stat in stats_mensuales:
        ws_mensual.append([
            stat['mes'].strftime('%B %Y'),
            stat['total'],
            f'${stat["ingresos"] or 0}',
            stat['canceladas']
        ])
    
    # Generar nombre de archivo
    fecha = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'estadisticas_complejo_{complejo.id}_{fecha}.xlsx'
    
    # Crear directorio si no existe
    os.makedirs('media/reportes', exist_ok=True)
    
    # Guardar archivo
    wb.save(f'media/reportes/{filename}')
    
    return filename
